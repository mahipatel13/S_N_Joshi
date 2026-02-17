from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify, send_file
import pymysql.cursors
from pymysql.cursors import DictCursor
import random
from datetime import date, datetime
import mysql.connector
import qrcode
import io
import subprocess
import pandas as pd
from fpdf import FPDF
from io import BytesIO
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from collections import defaultdict, OrderedDict
import uuid
import os
from io import BytesIO as _BytesIO
import openpyxl

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'sn_joshi_mess_secret_key_2024')

# Production flag for environment-aware settings
IS_PRODUCTION = os.environ.get('ENVIRONMENT', 'local').lower() == 'production'

# Store active sessions
ACTIVE_SESSIONS = {}

# Mess Name
MESS_NAME = "S. N. Joshi. Mess"

# Database configuration (environment-aware)
DB_CONFIG = {
    'host': os.environ.get('DB_HOST', 'localhost'),
    'port': int(os.environ.get('DB_PORT', 3306)),
    'user': os.environ.get('DB_USER', 'root'),
    'password': os.environ.get('DB_PASSWORD', ''),
    'db': os.environ.get('DB_NAME', 'track_serve'),
}

# Function to connect to MySQL
def get_db_connection():
    try:
        connection = pymysql.connect(
            host=DB_CONFIG['host'],
            port=DB_CONFIG['port'],
            user=DB_CONFIG['user'],
            password=DB_CONFIG['password'],
            db=DB_CONFIG['db'],
            cursorclass=pymysql.cursors.DictCursor
        )
        return connection
    except pymysql.MySQLError as e:
        print(f"Error connecting to MySQL: {e}")
        return None

@app.route('/')
def home():
    return redirect(url_for('adminlogin'))

@app.route('/health')
def health():
    return jsonify({"status": "ok"}), 200

@app.route('/favicon.ico')
def favicon():
    # Serve an existing static image so browser favicon requests don't 404.
    return redirect(url_for('static', filename='images/pari_1.png'))

@app.route('/adminlogin', methods=['GET', 'POST'])
def adminlogin():
    if request.method == 'POST':
        phone_no = request.form['phone_no'].strip()
        password = request.form['password']

        # Enforce 10-digit numeric phone number on server side
        if not phone_no.isdigit() or len(phone_no) != 10:
            return render_template('admin_login.html', error="Phone number must be exactly 10 digits", mess_name=MESS_NAME)

        connection = get_db_connection()
        if connection:
            try:
                cursor = connection.cursor()
                query = "SELECT * FROM admin WHERE phone_no = %s AND password = %s"
                cursor.execute(query, (phone_no, password))
                user = cursor.fetchone()

                if user:
                    # Check if this admin is already logged in elsewhere
                    admin_id = user['id'] if 'id' in user else phone_no
                    
                    # If already logged in, logout the previous session
                    if admin_id in ACTIVE_SESSIONS:
                        old_session_id = ACTIVE_SESSIONS[admin_id]
                        # Clear old session (if needed)
                        print(f"Logging out previous session for admin {admin_id}")
                    
                    # Create new session
                    session_id = str(uuid.uuid4())
                    ACTIVE_SESSIONS[admin_id] = session_id
                    
                    session['admin_id'] = admin_id
                    session['phone_no'] = phone_no
                    session['session_id'] = session_id
                    session['logged_in'] = True
                    
                    flash("Login successful! Welcome to S. N. Joshi. Mess", "success")
                    return redirect(url_for('admin_dashboard'))
                else:
                    return render_template('admin_login.html', error="Invalid phone number or password")

            except pymysql.MySQLError as e:
                print(f"Database error: {e}")
                return render_template('admin_login.html', error="Error while checking credentials")
            finally:
                cursor.close()
                connection.close()
        else:
            return render_template('admin_login.html', error="Failed to connect to the database")

    return render_template('admin_login.html', mess_name=MESS_NAME)

@app.before_request
def check_session():
    """Check if user session is valid"""
    if 'logged_in' in session:
        admin_id = session.get('admin_id')
        session_id = session.get('session_id')
        
        # Verify session is still active
        if admin_id not in ACTIVE_SESSIONS or ACTIVE_SESSIONS[admin_id] != session_id:
            session.clear()
            flash("Your session has been terminated. Please login again.", "warning")

@app.route('/admin_dashboard')
def admin_dashboard():
    if 'logged_in' not in session:
        return redirect(url_for('adminlogin'))
    return render_template("admin_dashboard.html", mess_name=MESS_NAME)

@app.route('/logout')
def logout():
    if 'admin_id' in session:
        admin_id = session['admin_id']
        if admin_id in ACTIVE_SESSIONS:
            del ACTIVE_SESSIONS[admin_id]
    
    session.clear()
    flash("You have been logged out successfully.", "success")
    return redirect(url_for('adminlogin'))

# Menu Management Routes
@app.route('/menu_dashboard')
def menu_dashboard():
    if 'logged_in' not in session:
        return redirect(url_for('adminlogin'))
    return render_template('menu_dashboard.html', mess_name=MESS_NAME)

@app.route('/add_menu')
def add_menu():
    if 'logged_in' not in session:
        return redirect(url_for('adminlogin'))
    return render_template('add_menu_dashboard.html', mess_name=MESS_NAME)

@app.route('/submit', methods=['POST'])
def submit():
    if 'logged_in' not in session:
        return redirect(url_for('adminlogin'))
        
    meal_type = request.form.get('meal_type')
    from_date = request.form.get("fromDate")
    to_date = request.form.get("toDate")

    print("From Date:", from_date)
    print("To Date:", to_date)

    # Fetch meal data for all days
    data = {
        'Monday': request.form.get('Monday'),
        'Tuesday': request.form.get('Tuesday'),
        'Wednesday': request.form.get('Wednesday'),
        'Thursday': request.form.get('Thursday'),
        'Friday': request.form.get('Friday'),
        'Saturday': request.form.get('Saturday'),
        'Sunday': request.form.get('Sunday')
    }

    # Identify the table to store meal data
    table_name = ""
    if meal_type == 'breakfast':
        table_name = "breakfast"
    elif meal_type == 'lunch':
        table_name = "lunch"
    elif meal_type == 'dinner':
        table_name = "dinner"

    if table_name:
        connection = get_db_connection()
        if connection:
            try:
                with connection.cursor() as cursor:
                    sql = f"""
                        INSERT INTO {table_name}
                        (Monday, Tuesday, Wednesday, Thursday, Friday, Saturday, Sunday, from_date, to_date)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """
                    cursor.execute(sql, tuple(data.values()) + (from_date, to_date))
                    connection.commit()
                    print(f"Data successfully inserted into {table_name}")
                    flash(f"Menu for {meal_type.upper()} added successfully!", "success")
                    return render_template("add_menu_dashboard.html", mess_name=MESS_NAME)
            except Exception as e:
                print("Database Error:", e)
                flash(f"Error: {str(e)}", "danger")
            finally:
                connection.close()

    return redirect(url_for('add_menu'))

def get_today_menu():
    try:
        conn = get_db_connection()
        if conn is None:
            return [{"Meal_Type": "Error", "Today_Menu": "MySQL Connection not available"}]

        cursor = conn.cursor()

        today = date.today().strftime('%Y-%m-%d')
        print(f"Today's Date: {today}")

        # Fetch the Weekday Name
        cursor.execute("SELECT DAYNAME(%s) AS weekday", (today,))
        result = cursor.fetchone()

        if not result or "weekday" not in result:
            print("Error: Could not retrieve weekday from MySQL")
            return [{"Meal_Type": "Error", "Today_Menu": "Could not retrieve weekday"}]

        weekday = result["weekday"]

        # Check if weekday column exists
        cursor.execute("SHOW COLUMNS FROM breakfast")
        breakfast_columns = [col["Field"] for col in cursor.fetchall()]

        if weekday not in breakfast_columns:
            print(f"Error: Column `{weekday}` does not exist in table.")
            return [{"Meal_Type": "Error", "Today_Menu": f"Column `{weekday}` not found"}]

        # Fetch Today's Menu
        query = f"""
        SELECT `{weekday}` AS Today_Menu, 'Breakfast' AS Meal_Type FROM breakfast 
        WHERE %s BETWEEN FROM_DATE AND TO_DATE
        UNION
        SELECT `{weekday}` AS Today_Menu, 'Lunch' AS Meal_Type FROM lunch 
        WHERE %s BETWEEN FROM_DATE AND TO_DATE
        UNION
        SELECT `{weekday}` AS Today_Menu, 'Dinner' AS Meal_Type FROM dinner 
        WHERE %s BETWEEN FROM_DATE AND TO_DATE;
        """
        cursor.execute(query, (today, today, today))
        result = cursor.fetchall()

        cursor.close()
        conn.close()

        return result if result else [{"Meal_Type": "No Data", "Today_Menu": "No menu available"}]
    
    except pymysql.MySQLError as e:
        print("SQL Error:", e)
        return [{"Meal_Type": "Error", "Today_Menu": str(e)}]

@app.route('/menu_show', methods=['GET', 'POST'])
def menu_show():
    if 'logged_in' not in session:
        return redirect(url_for('adminlogin'))
        
    today_menu = get_today_menu()
    return render_template("menu.html", menu=today_menu, mess_name=MESS_NAME)

# Function to Fetch Weekly Data
def fetch_weekly_data():
    """
    Fetch menu rows for the active week.
    If today's date has no matching week, fallback to the latest available week.
    """
    today_date = datetime.today().strftime('%Y-%m-%d')
    connection = get_db_connection()

    if not connection:
        return None, [], [], []

    try:
        with connection.cursor() as cursor:
            # Prefer current active week across all menu tables.
            cursor.execute(
                """
                SELECT FROM_DATE, TO_DATE
                FROM (
                    SELECT FROM_DATE, TO_DATE FROM breakfast
                    UNION ALL
                    SELECT FROM_DATE, TO_DATE FROM lunch
                    UNION ALL
                    SELECT FROM_DATE, TO_DATE FROM dinner
                ) AS weekly_ranges
                WHERE %s BETWEEN FROM_DATE AND TO_DATE
                ORDER BY FROM_DATE DESC
                LIMIT 1
                """,
                (today_date,)
            )
            date_range = cursor.fetchone()

            # If no current week exists, fallback to the latest inserted week.
            if not date_range:
                cursor.execute(
                    """
                    SELECT FROM_DATE, TO_DATE
                    FROM (
                        SELECT FROM_DATE, TO_DATE FROM breakfast
                        UNION ALL
                        SELECT FROM_DATE, TO_DATE FROM lunch
                        UNION ALL
                        SELECT FROM_DATE, TO_DATE FROM dinner
                    ) AS weekly_ranges
                    ORDER BY TO_DATE DESC, FROM_DATE DESC
                    LIMIT 1
                    """
                )
                date_range = cursor.fetchone()

            if not date_range:
                return None, [], [], []

            from_date = date_range.get("FROM_DATE") or date_range.get("from_date")
            to_date = date_range.get("TO_DATE") or date_range.get("to_date")

            if not from_date or not to_date:
                return None, [], [], []

            cursor.execute(
                "SELECT * FROM breakfast WHERE FROM_DATE = %s AND TO_DATE = %s ORDER BY id DESC",
                (from_date, to_date),
            )
            breakfast_data = cursor.fetchall()

            cursor.execute(
                "SELECT * FROM lunch WHERE FROM_DATE = %s AND TO_DATE = %s ORDER BY id DESC",
                (from_date, to_date),
            )
            lunch_data = cursor.fetchall()

            cursor.execute(
                "SELECT * FROM dinner WHERE FROM_DATE = %s AND TO_DATE = %s ORDER BY id DESC",
                (from_date, to_date),
            )
            dinner_data = cursor.fetchall()

        return date_range, breakfast_data, lunch_data, dinner_data
    except Exception as e:
        print("Database Error:", e)
        return None, [], [], []
    finally:
        connection.close()

@app.route('/weekly_data', methods=['GET'])
def weekly_data():
    if 'logged_in' not in session:
        return redirect(url_for('adminlogin'))
        
    date_range, breakfast, lunch, dinner = fetch_weekly_data()
    return render_template('weakly_menu.html', date_range=date_range, breakfast=breakfast, lunch=lunch, dinner=dinner, mess_name=MESS_NAME)

# PDF Generation Class
class CustomPDF(FPDF):
    def header(self):
        self.set_font("Arial", "B", 16)
        self.cell(0, 10, f"{MESS_NAME} - Weekly Menu", ln=True, align="C")
        self.ln(5)

# Function to Generate PDF
def generate_pdf_menu(from_date, meals):
    pdf = CustomPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=10)
    pdf.add_page()

    # Set Title
    pdf.set_font("Arial", "B", 12)
    pdf.cell(0, 10, f"Weekly Menu ({from_date})", ln=True, align="C")
    pdf.ln(5)

    # Column headers
    column_headers = ["Meal Type", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    column_widths = [40] + [35] * 7

    def draw_header_row():
        pdf.set_font("Arial", "B", 11)
        for i, header in enumerate(column_headers):
            pdf.cell(column_widths[i], 10, header, border=1, align="C")
        pdf.ln()

    def normalize_cell_text(value):
        if value is None:
            return ""
        text = str(value).replace("\r\n", "\n").replace("\r", "\n")
        # FPDF (core fonts) supports latin-1 only. Replace unsupported chars safely.
        text = text.encode("latin-1", "replace").decode("latin-1")
        cleaned_lines = []
        for line in text.split("\n"):
            compact = " ".join(line.split())
            if compact:
                cleaned_lines.append(compact)
        return "\n".join(cleaned_lines)

    def wrap_text_lines(text, width):
        text = normalize_cell_text(text)
        if not text:
            return [""]

        max_width = width - 2  # small inner padding
        wrapped = []
        for paragraph in text.split("\n"):
            words = paragraph.split()
            if not words:
                wrapped.append("")
                continue

            current_line = words[0]
            for word in words[1:]:
                candidate = f"{current_line} {word}"
                if pdf.get_string_width(candidate) <= max_width:
                    current_line = candidate
                else:
                    wrapped.append(current_line)
                    current_line = word
            wrapped.append(current_line)
        return wrapped if wrapped else [""]

    # Function to add meal rows dynamically with wrapped cell content
    def add_meal_row(meal_name, meal_data):
        row_values = [meal_name] + [meal_data.get(day, "") for day in column_headers[1:]]
        wrapped_cells = [
            wrap_text_lines(row_values[i], column_widths[i]) for i in range(len(row_values))
        ]

        line_height = 6
        row_height = max(len(lines) for lines in wrapped_cells) * line_height + 2

        # Add page break and redraw table header when needed
        if pdf.get_y() + row_height > (pdf.h - pdf.b_margin):
            pdf.add_page()
            draw_header_row()

        x_start = pdf.get_x()
        y_start = pdf.get_y()

        pdf.set_font("Arial", size=10)
        for i, lines in enumerate(wrapped_cells):
            x = x_start + sum(column_widths[:i])
            width = column_widths[i]

            # Draw cell border
            pdf.rect(x, y_start, width, row_height)

            # Draw wrapped text inside cell
            for idx, line in enumerate(lines):
                pdf.set_xy(x + 1, y_start + 1 + idx * line_height)
                align = "C" if i == 0 else "L"
                pdf.cell(width - 2, line_height, line, border=0, align=align)

        pdf.set_xy(x_start, y_start + row_height)

    draw_header_row()
    for meal_type, meal_data in meals.items():
        if not isinstance(meal_data, dict):
            print(f"Error: Expected dict for {meal_type}, but got {type(meal_data)}")
            continue
        add_meal_row(meal_type, meal_data)

    pdf_bytes = pdf.output(dest="S").encode("latin-1")
    return _BytesIO(pdf_bytes)

# Function to Extract Meal Data
def extract_meal_data(meal_data):
    if not meal_data:
        return {day: "" for day in ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]}

    if isinstance(meal_data[0], dict):
        meal_dict = meal_data[0]
    elif isinstance(meal_data[0], tuple):
        column_names = ["id", "from_date", "to_date", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
        meal_dict = dict(zip(column_names, meal_data[0]))
    else:
        print(f"Unexpected meal_data format: {meal_data[0]}")
        return {day: "" for day in ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]}

    return {day: meal_dict.get(day, "") for day in ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]}

# Route to Download Weekly Menu as PDF
@app.route('/download_menu_pdf', methods=['GET'])
def download_menu_pdf():
    if 'logged_in' not in session:
        return redirect(url_for('adminlogin'))

    try:
        date_range, breakfast, lunch, dinner = fetch_weekly_data()

        if not date_range:
            return jsonify({"error": "No data found for the current week"}), 404

        from_date = str(date_range.get("FROM_DATE") or date_range.get("from_date") or "").strip()
        safe_date = "".join(ch if ch.isalnum() or ch in ("-", "_") else "_" for ch in from_date) or "weekly_menu"

        meals = {
            "Breakfast": extract_meal_data(breakfast),
            "Lunch": extract_meal_data(lunch),
            "Dinner": extract_meal_data(dinner),
        }

        pdf_buffer = generate_pdf_menu(from_date or safe_date, meals)
        pdf_buffer.seek(0)
        filename = f"{safe_date}_Menu.pdf"
        return send_file(pdf_buffer, as_attachment=True, download_name=filename, mimetype="application/pdf")
    except Exception as e:
        app.logger.exception("PDF download failed: %s", e)
        return jsonify({"error": "Failed to generate PDF. Please try again."}), 500


@app.route('/download_menu_excel', methods=['GET'])
def download_menu_excel():
    if 'logged_in' not in session:
        return redirect(url_for('adminlogin'))

    date_range, breakfast, lunch, dinner = fetch_weekly_data()
    if not date_range:
        return jsonify({"error": "No menu data available for the current week. Please add weekly menu data first."}), 400

    # Prepare rows for Excel: one row per meal type
    def meal_to_row(name, meal_data):
        data = extract_meal_data(meal_data)
        row = { 'Meal Type': name }
        row.update(data)
        return row

    rows = [meal_to_row('Breakfast', breakfast), meal_to_row('Lunch', lunch), meal_to_row('Dinner', dinner)]
    df = pd.DataFrame(rows)

    output = _BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Weekly Menu')
    output.seek(0)
    from_date = str(date_range.get("FROM_DATE") or date_range.get("from_date") or "")
    filename = f"weekly_menu_{from_date}.xlsx" if from_date else "weekly_menu.xlsx"
    return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


def fetch_gv_rows():
    conn = get_db_connection()
    if not conn:
        return []

    try:
        with conn.cursor() as cursor:
            cursor.execute(
                """
                SELECT id, date_day, meal_type, menu_item, person, grocery, vegetable, khanabcha, khanaghata, remark
                FROM grocery_vegetable_management
                ORDER BY date_day DESC, id DESC
                """
            )
            return cursor.fetchall()
    except Exception as e:
        app.logger.exception("Failed to fetch grocery report data: %s", e)
        return []
    finally:
        conn.close()


@app.route('/download_gv_excel', methods=['GET'])
def download_gv_excel():
    if 'logged_in' not in session:
        return redirect(url_for('adminlogin'))

    rows = fetch_gv_rows()

    # Normalize rows into DataFrame
    if not rows:
        return jsonify({"error": "No grocery/vegetable data found"}), 404

    formatted_rows = []
    for row in rows:
        row_date = row.get("date_day")
        if isinstance(row_date, datetime):
            date_str = row_date.strftime("%Y-%m-%d")
            day_str = row_date.strftime("%A")
        elif isinstance(row_date, date):
            date_str = row_date.strftime("%Y-%m-%d")
            day_str = row_date.strftime("%A")
        else:
            date_str = str(row_date) if row_date else ""
            day_str = ""

        formatted_rows.append(
            {
                "Date": date_str,
                "Day": day_str,
                "Meal Type": row.get("meal_type", ""),
                "Menu": row.get("menu_item", ""),
                "Person": row.get("person", ""),
                "Grocery": row.get("grocery", ""),
                "Vegetable": row.get("vegetable", ""),
                "Khana Bacha": row.get("khanabcha", ""),
                "Khana Ghata": row.get("khanaghata", ""),
                "Remark": row.get("remark", ""),
            }
        )

    df = pd.DataFrame(formatted_rows)

    output = _BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Grocery_Vegetable_Report')
    output.seek(0)
    filename = f"grocery_vegetable_report_{datetime.today().strftime('%Y-%m-%d')}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/download_gv_pdf', methods=['GET'])
def download_gv_pdf():
    if 'logged_in' not in session:
        return redirect(url_for('adminlogin'))

    try:
        rows = fetch_gv_rows()
        if not rows:
            return jsonify({"error": "No grocery/vegetable data found"}), 404

        pdf = CustomPDF(orientation="L", unit="mm", format="A4")
        pdf.set_auto_page_break(auto=True, margin=10)
        pdf.add_page()
        pdf.set_font("Arial", "B", 12)
        pdf.cell(0, 10, "Grocery & Vegetable Report", ln=True, align="C")
        pdf.ln(2)

        headers = ["Date", "Day", "Meal", "Menu", "Person", "Grocery", "Vegetable", "Khana Bacha", "Khana Ghata", "Remark"]
        widths = [20, 16, 16, 52, 18, 30, 30, 24, 24, 24]

        def clean_text(value):
            text = "" if value is None else str(value)
            return text.encode("latin-1", "replace").decode("latin-1")

        def wrap_text(text, width):
            text = clean_text(text).replace("\r\n", "\n").replace("\r", "\n")
            max_width = width - 2
            lines = []
            for paragraph in text.split("\n"):
                words = paragraph.split()
                if not words:
                    lines.append("")
                    continue
                line = words[0]
                for word in words[1:]:
                    candidate = f"{line} {word}"
                    if pdf.get_string_width(candidate) <= max_width:
                        line = candidate
                    else:
                        lines.append(line)
                        line = word
                lines.append(line)
            return lines if lines else [""]

        def draw_header():
            pdf.set_font("Arial", "B", 9)
            for idx, header in enumerate(headers):
                pdf.cell(widths[idx], 8, header, border=1, align="C")
            pdf.ln()

        draw_header()
        pdf.set_font("Arial", size=8)
        line_height = 5

        for row in rows:
            row_date = row.get("date_day")
            if isinstance(row_date, (datetime, date)):
                date_str = row_date.strftime("%Y-%m-%d")
                day_str = row_date.strftime("%A")
            else:
                date_str = str(row_date) if row_date else ""
                day_str = ""

            values = [
                date_str,
                day_str,
                row.get("meal_type", ""),
                row.get("menu_item", ""),
                row.get("person", ""),
                row.get("grocery", ""),
                row.get("vegetable", ""),
                row.get("khanabcha", ""),
                row.get("khanaghata", ""),
                row.get("remark", ""),
            ]

            wrapped_cells = [wrap_text(values[i], widths[i]) for i in range(len(values))]
            row_height = max(len(cell) for cell in wrapped_cells) * line_height + 2

            if pdf.get_y() + row_height > (pdf.h - pdf.b_margin):
                pdf.add_page()
                draw_header()
                pdf.set_font("Arial", size=8)

            x_start = pdf.get_x()
            y_start = pdf.get_y()

            for idx, lines in enumerate(wrapped_cells):
                x = x_start + sum(widths[:idx])
                width = widths[idx]
                pdf.rect(x, y_start, width, row_height)

                for line_idx, line in enumerate(lines):
                    pdf.set_xy(x + 1, y_start + 1 + line_idx * line_height)
                    pdf.cell(width - 2, line_height, clean_text(line), border=0, align="L")

            pdf.set_xy(x_start, y_start + row_height)

        pdf_bytes = pdf.output(dest="S").encode("latin-1")
        output = _BytesIO(pdf_bytes)
        output.seek(0)
        filename = f"grocery_vegetable_report_{datetime.today().strftime('%Y-%m-%d')}.pdf"
        return send_file(output, as_attachment=True, download_name=filename, mimetype="application/pdf")
    except Exception as e:
        app.logger.exception("GV PDF download failed: %s", e)
        return jsonify({"error": "Failed to generate grocery report PDF. Please try again."}), 500

# Fetch all menu data for deletion
@app.route('/delete_menu', methods=['GET'])
def delete_menu():
    if 'logged_in' not in session:
        return redirect(url_for('adminlogin'))
        
    try:
        connection = get_db_connection()
        with connection.cursor() as cursor:
            breakfast_query = "SELECT * FROM breakfast"
            lunch_query = "SELECT * FROM lunch"
            dinner_query = "SELECT * FROM dinner"

            cursor.execute(breakfast_query)
            breakfast_data = cursor.fetchall()

            cursor.execute(lunch_query)
            lunch_data = cursor.fetchall()

            cursor.execute(dinner_query)
            dinner_data = cursor.fetchall()

        return render_template('delete_menu.html', breakfast=breakfast_data, lunch=lunch_data, dinner=dinner_data, mess_name=MESS_NAME)
    except Exception as e:
        print("Database Error:", e)
        return "Error fetching data"
    finally:
        if connection:
            connection.close()

# Handle menu deletion
@app.route('/delete_menu/<meal>/<int:menu_id>', methods=['POST'])
def delete_menu_item(meal, menu_id):
    if 'logged_in' not in session:
        return redirect(url_for('adminlogin'))
        
    try:
        connection = get_db_connection()
        with connection.cursor() as cursor:
            query = f"DELETE FROM {meal} WHERE id = %s"
            cursor.execute(query, (menu_id,))
            connection.commit()
        
        flash(f"Menu item deleted successfully", "success")
        return jsonify({"message": "Menu item deleted successfully"}), 200
    except Exception as e:
        print("Error deleting menu item:", e)
        return jsonify({"error": "Failed to delete"}), 500
    finally:
        if connection:
            connection.close()

# Grocery & Vegetable Management Routes
@app.route("/g_v_list", methods=["GET", "POST"])
def g_v_list():
    if 'logged_in' not in session:
        return redirect(url_for('adminlogin'))
        
    if request.method == "GET":
        meal_type = request.args.get("meal_type")
        menu_items = []
        if meal_type in ["breakfast", "lunch", "dinner"]:
            try:
                conn = get_db_connection()
                cursor = conn.cursor()
                today = datetime.now().date()
                weekday = today.strftime('%A')

                query = f"""
                    SELECT `{weekday}` AS day_value
                    FROM `{meal_type}`
                    WHERE `{weekday}` IS NOT NULL AND `{weekday}` != ''
                    AND from_date <= %s AND to_date >= %s
                """

                cursor.execute(query, (today, today))
                results = cursor.fetchall()

                for row in results:
                    if row['day_value']:
                        items = row['day_value'].split('\n')
                        menu_items.extend([item.strip() for item in items if item.strip()])

                cursor.close()
                conn.close()

            except pymysql.MySQLError as e:
                return f"Database error while fetching: {str(e)}", 500

        return render_template("g_v_list.html", meal_type=meal_type, menu_items=menu_items, mess_name=MESS_NAME)

    # POST processing
    meal_type = request.form.get("meal_type")
    persons = request.form.get("person")
    menu_items = request.form.getlist("menu_item[]")
    groceries = request.form.getlist("grocery[]")
    vegetables = request.form.getlist("vegetable[]")
    khanabchas = request.form.getlist("khanabcha[]")
    khanaghatas = request.form.getlist("khanaghata[]")
    remarks = request.form.getlist("remark[]")

    def format_readable_list(data_list):
        return "\n".join([f"{i+1}. {item.strip()}" for i, item in enumerate(data_list) if item.strip()])

    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("""
            INSERT INTO grocery_vegetable_management
            (date_day, meal_type, menu_item, person, grocery, vegetable, khanabcha, khanaghata, remark)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            datetime.now(),
            meal_type,
            format_readable_list(menu_items),
            persons,
            format_readable_list(groceries),
            format_readable_list(vegetables),
            format_readable_list(khanabchas),
            format_readable_list(khanaghatas),
            format_readable_list(remarks)
        ))

        conn.commit()
        cursor.close()
        conn.close()
        flash("Grocery and Vegetable data saved successfully!", "success")
    except pymysql.MySQLError as e:
        flash(f"Database error while inserting: {str(e)}", "danger")
        return redirect("/g_v_list")

    return redirect("/g_v_list")

@app.route("/g_v_report")
def g_v_report():
    if 'logged_in' not in session:
        return redirect(url_for('adminlogin'))

    rows = fetch_gv_rows()

    grouped_data = defaultdict(list)

    for row in rows:
        row_date = row.get('date_day')
        if isinstance(row_date, datetime):
            date_value = row_date.date()
            day_value = row_date.strftime("%A")
        elif isinstance(row_date, date):
            date_value = row_date
            day_value = row_date.strftime("%A")
        else:
            date_value = row_date
            day_value = ""

        meal_type = row['meal_type']
        key = (date_value, meal_type)
        grouped_data[key].append({
            'menu_item': row['menu_item'],
            'person': row['person'],
            'grocery': row['grocery'],
            'vegetable': row['vegetable'],
            'khanabcha': row['khanabcha'],
            'khanaghata': row['khanaghata'],
            'remark': row['remark'],
            'id': row.get('id') if isinstance(row, dict) else row['id'],
            'day': day_value,
            'date_str': date_value.strftime("%d-%m-%Y") if isinstance(date_value, date) else str(date_value)
        })

    sorted_grouped_data = OrderedDict(sorted(grouped_data.items(), key=lambda x: x[0][0], reverse=True))

    return render_template("g_v_report.html", grouped_data=sorted_grouped_data, mess_name=MESS_NAME)


@app.route("/delete_gv_row", methods=["POST"])
def delete_gv_row():
    if 'logged_in' not in session:
        return redirect(url_for('adminlogin'))
        
    # Prefer deleting by primary key `id` if provided (more reliable). Fall back
    # to the older menu_item LIKE behavior when `id` is not supplied.
    row_id = request.form.get("id")
    menu_item = request.form.get("menu_item")

    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        table_name = "grocery_vegetable_management"

        if row_id:
            query = f"DELETE FROM `{table_name}` WHERE id = %s"
            cursor.execute(query, (row_id,))
        elif menu_item:
            query = f"DELETE FROM `{table_name}` WHERE menu_item LIKE %s LIMIT 1"
            cursor.execute(query, (f"%{menu_item}%",))
        else:
            raise ValueError("No identifier provided for deletion")

        conn.commit()
        cursor.close()
        conn.close()
        flash("Row deleted successfully!", "success")
    except Exception as e:
        print("Deletion error:", e)
        flash(f"Error deleting row: {str(e)}", "danger")

    return redirect(url_for("g_v_report"))


@app.route('/g_v')
def g_v():
    if 'logged_in' not in session:
        return redirect(url_for('adminlogin'))
    return render_template('g_v.html', mess_name=MESS_NAME)


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 8080))
    debug = not IS_PRODUCTION  # Disable debug in production
    app.run(debug=debug, port=port, host='0.0.0.0')

